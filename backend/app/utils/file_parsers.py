"""File parsers for different document types"""
import io
import csv
import logging
from pathlib import Path
from typing import Dict, Any, List
import chardet

logger = logging.getLogger(__name__)

try:
    import openpyxl
    import pandas as pd
    EXCEL_AVAILABLE = True
except ImportError:
    EXCEL_AVAILABLE = False
    logger.warning("Excel parsing not available. Install openpyxl and pandas.")

try:
    import PyPDF2
    PDF_AVAILABLE = True
except ImportError:
    PDF_AVAILABLE = False
    logger.warning("PDF parsing not available. Install PyPDF2.")


class FileParser:
    """Base class for file parsers"""
    
    @staticmethod
    def parse(file_content: bytes, filename: str) -> Dict[str, Any]:
        """Parse file and return structured data"""
        raise NotImplementedError


class ExcelParser(FileParser):
    """Parser for Excel files (.xlsx, .xls)"""
    
    @staticmethod
    def parse(file_content: bytes, filename: str) -> Dict[str, Any]:
        """Parse Excel file and extract all sheets"""
        if not EXCEL_AVAILABLE:
            raise ValueError("Excel parsing not available. Install required dependencies.")
        
        try:
            # Read Excel file from bytes
            excel_file = io.BytesIO(file_content)
            workbook = openpyxl.load_workbook(excel_file, data_only=True)
            
            result = {
                "filename": filename,
                "type": "excel",
                "sheets": {}
            }
            
            for sheet_name in workbook.sheetnames:
                sheet = workbook[sheet_name]
                
                # Extract all data from sheet
                sheet_data = []
                for row in sheet.iter_rows(values_only=True):
                    # Skip empty rows
                    if any(cell is not None for cell in row):
                        sheet_data.append([str(cell) if cell is not None else "" for cell in row])
                
                result["sheets"][sheet_name] = {
                    "rows": len(sheet_data),
                    "data": sheet_data
                }
            
            # Also try pandas for better data extraction
            try:
                excel_file.seek(0)
                df_dict = pd.read_excel(excel_file, sheet_name=None, engine='openpyxl')
                
                for sheet_name, df in df_dict.items():
                    if sheet_name in result["sheets"]:
                        # Add pandas-parsed data for better structure
                        result["sheets"][sheet_name]["dataframe"] = df.to_dict('records')
                        result["sheets"][sheet_name]["columns"] = df.columns.tolist()
            except Exception as e:
                logger.warning(f"Pandas parsing failed for {filename}: {e}")
            
            logger.info(f"Successfully parsed Excel file: {filename} ({len(result['sheets'])} sheets)")
            return result
            
        except Exception as e:
            logger.error(f"Error parsing Excel file {filename}: {e}")
            raise ValueError(f"Failed to parse Excel file: {str(e)}")


class PDFParser(FileParser):
    """Parser for PDF files"""
    
    @staticmethod
    def parse(file_content: bytes, filename: str) -> Dict[str, Any]:
        """Parse PDF and extract text"""
        if not PDF_AVAILABLE:
            raise ValueError("PDF parsing not available. Install PyPDF2.")
        
        try:
            pdf_file = io.BytesIO(file_content)
            pdf_reader = PyPDF2.PdfReader(pdf_file)
            
            result = {
                "filename": filename,
                "type": "pdf",
                "pages": len(pdf_reader.pages),
                "content": []
            }
            
            for page_num, page in enumerate(pdf_reader.pages):
                text = page.extract_text()
                if text.strip():
                    result["content"].append({
                        "page": page_num + 1,
                        "text": text
                    })
            
            # Concatenate all text
            result["full_text"] = "\n\n".join([p["text"] for p in result["content"]])
            
            logger.info(f"Successfully parsed PDF file: {filename} ({result['pages']} pages)")
            return result
            
        except Exception as e:
            logger.error(f"Error parsing PDF file {filename}: {e}")
            raise ValueError(f"Failed to parse PDF file: {str(e)}")


class TextParser(FileParser):
    """Parser for text files (.txt, .md)"""
    
    @staticmethod
    def parse(file_content: bytes, filename: str) -> Dict[str, Any]:
        """Parse text file"""
        try:
            # Detect encoding
            detected = chardet.detect(file_content)
            encoding = detected['encoding'] or 'utf-8'
            
            # Decode content
            text = file_content.decode(encoding)
            
            result = {
                "filename": filename,
                "type": "text",
                "encoding": encoding,
                "lines": len(text.split('\n')),
                "content": text
            }
            
            # If markdown, try to extract sections
            if filename.endswith('.md'):
                result["type"] = "markdown"
                result["sections"] = TextParser._extract_markdown_sections(text)
            
            logger.info(f"Successfully parsed text file: {filename} ({result['lines']} lines)")
            return result
            
        except Exception as e:
            logger.error(f"Error parsing text file {filename}: {e}")
            raise ValueError(f"Failed to parse text file: {str(e)}")
    
    @staticmethod
    def _extract_markdown_sections(text: str) -> List[Dict[str, str]]:
        """Extract sections from markdown by headers"""
        sections = []
        current_section = None
        current_content = []
        
        for line in text.split('\n'):
            if line.startswith('#'):
                # Save previous section
                if current_section:
                    sections.append({
                        "title": current_section,
                        "content": '\n'.join(current_content).strip()
                    })
                
                # Start new section
                current_section = line.lstrip('#').strip()
                current_content = []
            else:
                current_content.append(line)
        
        # Save last section
        if current_section:
            sections.append({
                "title": current_section,
                "content": '\n'.join(current_content).strip()
            })
        
        return sections


class CSVParser(FileParser):
    """Parser for CSV files"""
    
    @staticmethod
    def parse(file_content: bytes, filename: str) -> Dict[str, Any]:
        """Parse CSV file"""
        try:
            # Detect encoding
            detected = chardet.detect(file_content)
            encoding = detected['encoding'] or 'utf-8'
            
            # Decode content
            text = file_content.decode(encoding)
            
            # Try different delimiters
            dialect = csv.Sniffer().sniff(text[:1024])
            
            # Parse CSV
            csv_reader = csv.DictReader(io.StringIO(text), dialect=dialect)
            rows = list(csv_reader)
            
            result = {
                "filename": filename,
                "type": "csv",
                "encoding": encoding,
                "rows": len(rows),
                "columns": csv_reader.fieldnames,
                "data": rows
            }
            
            logger.info(f"Successfully parsed CSV file: {filename} ({result['rows']} rows)")
            return result
            
        except Exception as e:
            logger.error(f"Error parsing CSV file {filename}: {e}")
            raise ValueError(f"Failed to parse CSV file: {str(e)}")


def parse_file(file_content: bytes, filename: str) -> Dict[str, Any]:
    """Parse file based on extension"""
    ext = Path(filename).suffix.lower()
    
    parsers = {
        '.xlsx': ExcelParser,
        '.xls': ExcelParser,
        '.pdf': PDFParser,
        '.txt': TextParser,
        '.md': TextParser,
        '.csv': CSVParser,
    }
    
    parser = parsers.get(ext)
    if not parser:
        raise ValueError(f"Unsupported file type: {ext}")
    
    return parser.parse(file_content, filename)
